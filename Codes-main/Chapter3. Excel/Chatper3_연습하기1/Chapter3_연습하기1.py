import openpyxl as op
from openpyxl import Workbook

#step1. 엑셀 Workbook 및 Worksheet 객체 생성
path = r"C:\Users\Yongbeom Jeong\Desktop\VS CODE\Project\사장님몰래하는파이썬업무자동화\연습하기\Chatper3_연습하기1/Chapter3_연습하기1_sample.xlsx"
wb = op.load_workbook(path)
ws = wb.active

#step2. 데이터를 저장하기 위한 빈 리스트 생성
category_list =[] #종류
data_list =[] #data

#step3. rows 속성 사용하여 A열, B열을 각각 리스트에 추가하기
for r in ws.rows:
    category_list.append(r[0].value) #튜플의 0번째 요소값
    data_list.append(r[1].value) #튜플이 1번째 요소값
  
print("category list 출력 : ", category_list)
print("data_list 출력 : ", data_list)

#step4. split 함수를 통한 분리 결과를 리스트에 저장
i=0 #종류를 표시하는 인덱스
result_list = [] #분리 결과를 저장할 list

#step4. split 함수를 통한 분리 결과를 리스트에 저장
for data in data_list:
    temp_text = data.split(',') #콤마 기준으로 분리
    print(temp_text)
    for word in temp_text:
        temp_tuple = (category_list[i], word)
        result_list.append(temp_tuple)
print(result_list)
i=i+1

#step5. 엑셀에 새로 쓰기
ws_result = wb.create_sheet("결과")

row_num=1
for one in result_list:
    ws_result.cell(row=row_num, column=1).value = one[0]
    ws_result.cell(row=row_num, column=2).value = one[1]
    row_num=row_num+1

wb.save("분리결과.xlsx")
wb.close


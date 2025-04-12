#step1. 엑셀 Workbook 및 Worksheet 객체 생성
import openpyxl as op
from  openpyxl.styles  import  Alignment


path = r"C:\Users\Yongbeom Jeong\Desktop\VS CODE\Project\사장님몰래하는파이썬업무자동화\연습하기\Chatper3_연습하기2\Chapter3_연습하기2_sample.xlsx"
wb = op.load_workbook(path)
ws = wb.active

#step2 : 연락처 텍스트 가공 및 입력
temp_list = [] #빈 리스트 생성
row_num = ws.max_row  #최대 행값 저장
for num in range(2, row_num+1): #2부터인 이유는 제목행 제외
    text = ws.cell(row=num, column=4).value #원본 D열 연락처를 저장
    text = text.split('/') #읽어온 연락처를 슬래쉬('/') 기준으로 구분
    cell_phone = text[0][-13:] #문자열 슬라이싱으로 연락처 읽어옴
   
    #예외처리 구문 : 사무실 번호가 없는 경우에는 "N/A" 처리 한다.
    try:
        office_phone = text[1][-13:]
    except:
        office_phone = "N/A"
   
    #셀 주소로 접근하여 각각 연락처를 구분하여 입력한다.(D,E열)
    ws["D"+str(num)].value = cell_phone
    ws["E"+str(num)].value = office_phone

#step3 : 양식 가공 및 저장
ws["D1"].value = "핸드폰" #제목행 입력
ws["E1"].value = "사무실" #제목행 입력
ws.column_dimensions['D'].width = 30 #C열의 너비 50으로
ws.column_dimensions['E'].width = 30 #C열의 너비 50으로
for r in ws.rows:
    r[3].alignment = Alignment(horizontal = 'center', vertical='center')
    r[4].alignment = Alignment(horizontal = 'center', vertical='center')


#저장 및 Workbook 닫기
wb.save("result.xlsx")
wb.close()

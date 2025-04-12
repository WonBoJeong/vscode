# step1.필요한 패키지와 모듈 불러오기
from selenium import webdriver #셀레니움 웹드라이버 제어 모듈
from selenium.webdriver.common.by import By  #HTML 인덱싱을 위한 클래스
from selenium.webdriver.common.keys import Keys # 크롬 드라이버로 원하는 키를 입력하기 위한 클래스
import time #시간 지연을 위한 모듈
import pandas as pd # 크롤링한 데이터를 표 형태로 정리하기 위한 패키지
import json # 프로그램 외부에서 변수를 변경하기 위한 모듈 (ip 주소)
import datetime # 현재 날짜를 파악하기 위한 모듈
import os # 폴더 및 파일의 경로를 생성하기 위한 모듈
import win32com.client as win32 # 엑셀을 꾸며주기 위한 모듈
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.borders import Border, Side

# step2.함수로 각 기능을 모듈화

# func1.hsmoa 사이트에서 어제 방송 시간,방송명 크롤링 & 방송녹화 파일 링크 추가하여 엑셀 파일로 출력
def hsmoa_crawling_info_to_excel():
    # 오늘 날짜 파악
    now = time.localtime()
    yesterday = str(now.tm_year) + str(now.tm_mon).zfill(2) + str(now.tm_mday-1).zfill(2)

    #크롬 드라이버 실행
    driver = webdriver.Chrome()
    # 홈쇼핑 사이트 접속
    url = "https://hsmoa.com/?date="+yesterday+"&site=&cate="
    driver.get(url)
    time.sleep(1)

    # 채널명, 채널클릭용 class명, mp4저장시 사용할 이름 담을 리스트 선언
    shopping_channels = ['신세계라이브쇼핑', 'KT알파쇼핑', 'W쇼핑', '쇼핑엔티', 'SK스토아', '현대홈쇼핑PLUS', 'CJ온스타일PLUS', '롯데OneTV', 'GSMYSHOP', 'NSSHOPPLUS']
    channels = ['bt-channel.ssgshop', 'bt-channel.kshop', 'bt-channel.wshop', 'bt-channel.shopnt', 'bt-channel.bshop', 'bt-channel.hmallplus', 'bt-channel.cjmallplus', 'bt-channel.lotteonetv', 'bt-channel.gsmyshop', 'bt-channel.nsmallplus']
    channels_save_mp4 = ['SSG', 'KT', 'W', 'ST', 'SK', 'Hmall', 'CJ', 'Lotte', 'GS', 'NS']

    # 크롤링한 결과물을 담을 리스트 선언
    list_airtime = []
    list_pgm= []
    list_mp4 = []

    # 엑셀 파일에 채널별 크롤링 정보를 시트별로 저장하기 위하여 크롤링초반에 엑셀 객체 생성 
    writer = pd.ExcelWriter(f'hsmoa 방송 편성표_{yesterday}.xlsx', engine='xlsxwriter')

    # 10개의 채널을 동일한 규칙으로 반복 크롤링하기 위한 for문
    for num, channel in enumerate(channels):

        # 미리 저장해뒀던 각 채널별 class명을 이용해서 채널을 하나씩 클릭
        driver.find_element(By.CLASS_NAME, channel).click()
        time.sleep(2)

        # 크롤링 할 내용들을 웹브라우저가 인식할 수 있도록 스크롤 다운
        driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.END)
        time.sleep(2)
        driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_UP)
        time.sleep(1)
        
        # 편성시간
        airtime = driver.find_elements(By.CLASS_NAME, "font-12.c-midgray")
        list_airtime = [i.text for i in airtime if i.text != ""]

        # PGM
        title = driver.find_elements(By.CLASS_NAME, "font-15")
        list_pgm = [i.text for i in title if i.text != ""]
        
        # 방송바로보기 링크 (func2 여기서 사용)
        list_mp4 = [f"http://10.10.5.0.44:8080/files/{shopping_channels[num]}/{now.tm_mon}월/{now.tm_mday-1}일/{channels_save_mp4[num]}_{match_time_slot(airtime)}.mp4" for airtime in list_airtime]

        # list 취합한 후 DataFrame 생성
        list_sum = list(zip(list_airtime, list_pgm, list_mp4))
        col = ['방송시간', 'PGM', '영상 바로보기']
        df = pd.DataFrame(list_sum, columns=col)
        
        # Save DataFrame to Excel sheet with shopping channel name as sheet name
        sheet_name = shopping_channels[num]
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 엑셀 파일 저장 및 크롤러 종료
    writer._save()
    driver.quit()


# func2.방송시간(air_time)에 따라 매칭되는 영상 파일명 결정 함수 (time_str은 0시 37분 ~ 1시 38분와 같은 형식)
def match_time_slot(time_str):
    start_hour = int(time_str.split()[0].split('시')[0]) #0
    start_minute = int(time_str.split()[1].split('분')[0]) #37
    start_time = datetime.time(start_hour, start_minute) # 00:37

    if start_time >= datetime.time(0, 0) and start_time < datetime.time(2, 0): #00:00 ~ 01:59
        return '00시-02시'
    elif start_time >= datetime.time(2, 0) and start_time < datetime.time(4, 0):
        return '02시-04시'
    elif start_time >= datetime.time(4, 0) and start_time < datetime.time(6, 0):
        return '04시-06시'
    elif start_time >= datetime.time(6, 0) and start_time < datetime.time(8, 0):
        return '06시-08시'
    elif start_time >= datetime.time(8, 0) and start_time < datetime.time(10, 0):
        return '08시-10시'
    elif start_time >= datetime.time(10, 0) and start_time < datetime.time(12, 0):
        return '10시-12시'
    elif start_time >= datetime.time(12, 0) and start_time < datetime.time(14, 0):
        return '12시-14시'
    elif start_time >= datetime.time(14, 0) and start_time < datetime.time(16, 0):
        return '14시-16시'
    elif start_time >= datetime.time(16, 0) and start_time < datetime.time(18, 0):
        return '16시-18시'
    elif start_time >= datetime.time(18, 0) and start_time < datetime.time(20, 0):
        return '18시-20시'
    elif start_time >= datetime.time(20, 0) and start_time < datetime.time(22, 0):
        return '20시-22시'
    elif start_time >= datetime.time(22, 0) and start_time <= datetime.time(23, 59):
        return '22시-24시'
    elif start_time >= datetime.time(0, 0) and start_time < datetime.time(2, 0):
        return '22시-24시'  # 다음날 00시부터 02시까지는 전날 22시~24시에 포함됨
    else:
        return None

# func3.크롤링 된 엑셀을 정해진 양식으로 예쁘게 변환하는 함수
def format_excel_file(file_path):
    
    from openpyxl.styles import Font
    
    # 엑셀 파일 로드
    wb = load_workbook(filename=file_path)

    # 글자 크기 설정
    font = Font(size=8)
    
    # Border 객체 생성
    border = Border(left=Side(style='none'),
                    right=Side(style='none'),
                    top=Side(style='none'),
                    bottom=Side(style='none'))

    for ws in wb.worksheets:
        # 모든 셀에 Border 적용
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.font = font
                cell.hyperlink = None

        # 첫번째 행 높이 설정
        ws.row_dimensions[1].height = 30

        # 열 너비 설정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 50

        # 첫번째 행의 스타일 설정
        first_row = ws[1]

        for cell in first_row:
            # 셀 배경색 설정
            cell.fill = PatternFill(start_color='FFF2F4', end_color='FFF2F4', fill_type='solid')
            # 셀 가운데 정렬
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 수정된 엑셀 파일 저장
    wb.save(filename=file_path)

# func4.최종 엑셀 파일을 시트마다 하나의 pdf 파일로 저장
def excel_to_pdf(excel_file_path, pdf_output_path):
    
    # 현재 날짜 및 시간 정보 가져오기
    now = time.localtime()

    # PDF 출력 경로가 없으면 생성
    if not os.path.exists(pdf_output_path):
        os.makedirs(pdf_output_path)

    # 엑셀 파일 불러오기
    wb = load_workbook(filename=excel_file_path)

    # Excel을 PDF 프린터로 출력
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False

    # 모든 시트에 대해 반복하며 PDF로 출력
    for sheet_name in wb.sheetnames:

        # PDF 파일명 설정
        pdf_file_name = pdf_output_path + sheet_name + f"_{now.tm_mon}월"+f"{now.tm_mday-1}일" + ".pdf"

        # PageSetup 객체 가져오기
        ws = excel.Workbooks.Open(excel_file_path).Worksheets(sheet_name)
        ws.PageSetup.Orientation = 2  # 1: 가로, 2: 세로
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = False
        ws.PageSetup.TopMargin = 0.5
        ws.PageSetup.BottomMargin = 0.5
        ws.PageSetup.RightMargin = 0.5
        ws.PageSetup.LeftMargin = 0.75

        # ExportAsFixedFormat 함수를 이용해서 엑셀 시트를 PDF 형식으로 내보내기
        ws.ExportAsFixedFormat(0, pdf_file_name)

    # Excel 종료
    excel.Quit()

def main():
    # hsmoa 크롤링 및 엑셀 파일에 저장
    hsmoa_crawling_info_to_excel() #func1 (func2는 func1 안에서 작동)

    # 오늘 날짜 파악
    now = time.localtime()
    yesterday = str(now.tm_year) + str(now.tm_mon).zfill(2) + str(now.tm_mday-1).zfill(2)

    # 저장한 엑셀 파일을 포맷에 맞게 예쁘게 수정
    format_excel_file(f'hsmoa 방송 편성표_{yesterday}.xlsx') #func3

    # 엑셀 파일의 각 시트를 pdf 파일로 변환하여 하나의 폴더에 저장
    excel_file_path = os.getcwd() + f'\\hsmoa 방송 편성표_{yesterday}.xlsx'
    pdf_output_path = os.getcwd() + '\\hsmoa_daily_pdf\\'
    print("PDF 생성 중...")
    excel_to_pdf(excel_file_path, pdf_output_path) #func4
    time.sleep(1)

    # 변환이 완료되면 엑셀 파일은 삭제
    try:
        os.remove(excel_file_path)
    except:
        print("Fail to delete Excel file")
        
main()
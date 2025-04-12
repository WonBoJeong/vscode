#엑셀 데이터를 자동으로 시트로 분류하기
import openpyxl as op

#함수1 : 엑셀 각 행 데이터를 저장해서 이차원 리스트 만들기
def makeExcelDataList(ws:object, offset_num:int=None):
    result=[] #빈 리스트 생성(결과 저장)
    for one_line in ws.rows: # 각 행을 for loop문 진행
        list_1row =[one.value for one in one_line]  # 리스트 컴프리헨션 엑셀 1행 저장
        result.append(list_1row) # 결과리스트에 추가 
    if offset_num is not None: #제목행 제외조건
         result = result[offset_num:]       
    return result

#함수2 : 열번호를 받아서 해당 열의 데이터를 카테고리화
def makeCategoryList(excel_list:list, col_num:int):
    data_list=[] #빈 결과 리스트 생성
    for one_line in excel_list: #엑셀 리스트를 for 반복문 진행
        data_list.append(one_line[col_num-1]) #4번열 일 경우 배열에서는 3번 열 참조
    category_set = set(data_list) #중복을 제거하기 위해 set 사용
    return list(category_set) #결과를 리스트로 리턴

#함수3 : Category list와 열번호를 입력받아서 list 별 분류하기 
def classficationData(excel_list:list, col_num:int, cat_list:list):
    result_dict={} # 결과 저장을 위한 빈 딕셔너리 생성
    for cat in cat_list: #카테고리 [A,B]를 반복
        save_list = [] # 각 카테고리마다 분류 된 데이터 저장할 리스트
        for one_line in excel_list: # 엑셀의 각 행을 반복
            if one_line[col_num-1] == cat: # 분류를 원하는 열값과 카테고리 분류값 비교
                save_list.append(one_line) # 같으면 배열에 저장
        result_dict[cat] = save_list # 결과 딕셔너리 저장(해당 카테고리에 대한 각 엑셀 행)
    return result_dict
        
        
#함수4: 새로운 분류 데이터를 엑셀로 저장
def writeCategoryData(cat_dict:dict, save_path:str, row_offset=1):
    wb = op.Workbook() # 새로운 Workboom 객체 생성
    for cat_name, data_1row in cat_dict.items(): #딕셔너리 반복문
        ws = wb.create_sheet(cat_name)
        row_idx = row_offset #입력을 시작할 엑셀 행(기본값 1)
        for one_data in data_1row:
            ws.cell(row=row_idx, column=1).value = one_data[0] # 제품명
            ws.cell(row=row_idx, column=2).value = one_data[1] # Serial number
            ws.cell(row=row_idx, column=3).value = one_data[2] # 날짜
            ws.cell(row=row_idx, column=4).value = one_data[3] # 대리점
            row_idx+=1
    wb.save(save_path+'/'+'result.xlsx') #결과 파일 저장
    wb.close() # 엑셀 객체 닫기

if __name__ == '__main__':
    path = r"./출하현황"
    file_name = "6월_제품_출하현황.xlsx"
    wb = op.load_workbook(path+"/"+file_name)
    ws = wb.active
    data_list = makeExcelDataList(ws, offset_num=1) #함수호출    
    cat_list = makeCategoryList(data_list, 4)
    classification = classficationData(data_list, 4, cat_list)
    writeCategoryData(classification, path)